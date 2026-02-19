import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill
from .logic import analyze_links, extract_basic_auth


def run():

    st.title("🔍 Link Behavior Audit Tool")

    st.markdown(
        """
This app scans a webpage for all **links** and checks:
- Same Tab / New Tab
- Internal / External
- Expected behavior
- HTTP health
"""
    )

    url = st.text_input("Enter Website URL", "https://www.broadridge.com/")

    st.info(
        "🔐 DEV access:\n\n"
        "`https://username:password@www-dev.example.com`"
    )

    if st.button("Run Audit"):

        with st.spinner("Analyzing links..."):
            clean_url, username, password = extract_basic_auth(url)
            results, error = analyze_links(clean_url, username, password)

        if error:
            st.error(error)

        elif results:
            df = pd.DataFrame(results)
            st.success(f"Found {len(df)} links.")

            def highlight_row(row):
                if row.get("Link Health") in ["Client Error", "Server Error", "Unreachable"]:
                    return ["background-color: #f8d7da"] * len(row)
                elif row["Expected?"] == "✘":
                    return ["background-color: #fff3cd"] * len(row)
                else:
                    return ["background-color: #d4edda"] * len(row)

            styled_df = df.style.apply(highlight_row, axis=1)

            st.dataframe(styled_df, use_container_width=True)

            # Excel export
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="AuditResults")
                ws = writer.sheets["AuditResults"]

                for r in range(2, len(df) + 2):
                    expected_value = df.iloc[r - 2]["Expected?"]
                    fill = PatternFill(
                        start_color="D4EDDA" if expected_value == "✔" else "F8D7DA",
                        end_color="D4EDDA" if expected_value == "✔" else "F8D7DA",
                        fill_type="solid",
                    )
                    for c in range(1, len(df.columns) + 1):
                        ws.cell(row=r, column=c).fill = fill

            st.download_button(
                "⬇️ Download results as Excel",
                data=output.getvalue(),
                file_name="link_audit_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        else:
            st.warning("No links found.")