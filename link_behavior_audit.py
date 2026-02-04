import requests
from urllib.parse import urlparse, urljoin
import streamlit as st
import pandas as pd
import io

# ✅ Must be the first Streamlit call
st.set_page_config(page_title="Link Behavior Audit", layout="wide")

from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright

# -------------------------
# Configuration
# -------------------------

SPECIAL_EXTERNAL_PATHS = ["/fr/", "/de/", "/jp/", "/next"]

IGNORE_SELECTORS = [
    "footer#footer-section",
    "nav#main-nav",
    "div.promo-bar",
    "div.contact-us__bottom",
    "div.ot-sdk-row",
    "div#onetrust-consent-sdk",
    "div#onetrust-group-container",
    "div.ot-pc-footer-logo",
    "a.ot-cookie-policy-link",
    "div.recaptcha-disclaimer",
    "div.recaptcha-disclaimer.reducefont",
    "a.skip-link[href='#main-content']",
]

IGNORE_HREF_PATTERNS = [
    "https://policies.google.com/privacy",
    "https://policies.google.com/terms",
    "javascript:void(0)",
    "/contact-us",
]

SOCIAL_DOMAINS = ["twitter.com", "facebook.com", "linkedin.com"]

# -------------------------
# Helpers
# -------------------------

def extract_basic_auth(url: str):
    """
    Supports:
    https://username:password@www-dev.example.com
    """
    parsed = urlparse(url)

    if parsed.username and parsed.password:
        clean_url = parsed._replace(netloc=parsed.hostname).geturl()
        return clean_url, parsed.username, parsed.password

    return url, "", ""


def check_link_status(url, timeout=10):
    try:
        response = requests.head(url, allow_redirects=True, timeout=timeout)

        if response.status_code >= 400:
            response = requests.get(url, allow_redirects=True, timeout=timeout)

        code = response.status_code

        if 200 <= code < 300:
            return code, "OK"
        elif 300 <= code < 400:
            return code, "Redirect"
        elif 400 <= code < 500:
            return code, "Client Error"
        elif 500 <= code < 600:
            return code, "Server Error"
        else:
            return code, "Unknown"

    except Exception:
        return "Error", "Unreachable"


# -------------------------
# Core Logic
# -------------------------

def analyze_links(page_url, username="", password=""):
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )

            context = browser.new_context(
                http_credentials={
                    "username": username,
                    "password": password,
                }
                if username and password
                else None
            )

            page = context.new_page()
            page.goto(page_url, wait_until="networkidle", timeout=30000)
            page.wait_for_selector("a", timeout=10000)

            html = page.content()

            context.close()
            browser.close()

    except Exception as e:
        return [], f"Error fetching page: {e}"

    soup = BeautifulSoup(html, "html.parser")

    for selector in IGNORE_SELECTORS:
        for tag in soup.select(selector):
            tag.decompose()

    links = soup.find_all("a", href=True)
    base_domain = urlparse(page_url).netloc

    results = []

    for link in links:
        href = link.get("href")
        if not href:
            continue

        if href.startswith(("tel:", "mailto:", "#main-content")):
            continue

        if any(pattern in href for pattern in IGNORE_HREF_PATTERNS):
            continue

        absolute_url = urljoin(page_url, href)
        parsed_url = urlparse(absolute_url)

        link_text = (
            link.get_text(strip=True)
            or link.get("aria-label")
            or link.get("title")
            or absolute_url
        )

        is_external = parsed_url.netloc and parsed_url.netloc != base_domain
        for sp in SPECIAL_EXTERNAL_PATHS:
            if parsed_url.path.startswith(sp) and parsed_url.netloc == base_domain:
                is_external = True
                break

        opens_in = "New Tab" if link.get("target") == "_blank" else "Same Tab"

        if any(domain in href for domain in SOCIAL_DOMAINS):
            opens_in = "New Tab"

        doc_extensions = (
            ".pdf",
            ".doc",
            ".docx",
            ".xls",
            ".xlsx",
            ".csv",
            ".ppt",
            ".pptx",
        )
        is_document = absolute_url.lower().endswith(doc_extensions)

        if is_document:
            is_external = True
            expected = "✔" if opens_in == "New Tab" else "✘"
            reason = (
                "Document: must open External in New Tab"
                if expected == "✔"
                else "Document opened wrong"
            )
        else:
            expected = (
                "✔"
                if (is_external and opens_in == "New Tab")
                or (not is_external and opens_in == "Same Tab")
                else "✘"
            )

            if is_external:
                reason = "External OK" if expected == "✔" else "External should open New Tab"
            else:
                reason = "Internal OK" if expected == "✔" else "Internal should open Same Tab"

        status_code, link_health = check_link_status(absolute_url)

        results.append(
            {
                "Link Text": link_text,
                "Opens In": opens_in,
                "Internal/External": "External" if is_external else "Internal",
                "HTTP Status": status_code,
                "Link Health": link_health,
                "Expected?": expected,
                "Reason": reason,
            }
        )

    return results, None


# -------------------------
# Streamlit UI
# -------------------------

st.title("🔍 Link Behavior Audit Tool")

st.markdown(
    """
This app scans a webpage for all **links** and checks:
- Whether they open in **Same Tab** or **New Tab**
- Whether they are **Internal** or **External**
- Whether the behavior matches the **Expected Rule**
- Whether the link is **reachable or broken** (HTTP status)

**Rules**
- Internal → Same Tab  
- External → New Tab  
- Documents → Always External + New Tab  
- 4xx / 5xx → Broken links
"""
)

url = st.text_input("Enter Website URL", "https://www.broadridge.com/")

st.info(
    "🔐 **DEV environment access**\n\n"
    "If the page is protected by Basic Authentication, use:\n\n"
    "`https://username:password@www-dev.broadridge.com`\n\n"
    "For public or PROD URLs, use the normal URL."
)

if st.button("Run Audit"):
    with st.spinner("Analyzing links..."):
        clean_url, username, password = extract_basic_auth(url)
        results, error = analyze_links(clean_url, username, password)

    if error:
        st.error(error)

    elif results:
        df = pd.DataFrame(results)
        st.success(f"Found {len(df)} links.")

        def highlight_row(row):
            if row.get("Link Health") in ["Client Error", "Server Error", "Unreachable"]:
                return ["background-color: #f8d7da"] * len(row)
            elif row["Expected?"] == "✘":
                return ["background-color: #fff3cd"] * len(row)
            else:
                return ["background-color: #d4edda"] * len(row)

        df = df.reset_index(drop=True)
        styled_df = df.style.apply(highlight_row, axis=1)

        html_table = styled_df.hide(axis="index").to_html(escape=False)

        st.components.v1.html(
            f"""
            <div style="max-height:450px; overflow-y:auto; border:1px solid #ddd;">
                {html_table}
            </div>
            """,
            height=500,
            scrolling=False,
        )

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="AuditResults")
            ws = writer.sheets["AuditResults"]
            for r in range(2, len(df) + 2):
                expected_value = df.iloc[r - 2]["Expected?"]
                fill = PatternFill(
                    start_color="D4EDDA" if expected_value == "✔" else "F8D7DA",
                    end_color="D4EDDA" if expected_value == "✔" else "F8D7DA",
                    fill_type="solid",
                )
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=r, column=c).fill = fill

        st.download_button(
            "⬇️ Download results as Excel",
            data=output.getvalue(),
            file_name="link_audit_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    else:
        st.warning("No links found on this page.")
